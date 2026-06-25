"""
Testes unitários — exportação (xlsx/csv).

Cobre injeção CSV, hyperlinks Excel e fallback para registros antigos.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

# Stub de dependências pesadas que não estão disponíveis no ambiente de testes
from unittest.mock import MagicMock
for _mod in ("pytesseract", "cv2", "PIL", "PIL.Image"):
    if _mod not in sys.modules:
        sys.modules[_mod] = MagicMock()

import io
import json
from datetime import datetime

import openpyxl
import pytest

from services.verificacao import _sanitizar_csv, gerar_csv, gerar_xlsx, _NAO_DISPONIVEL


# ── _sanitizar_csv ────────────────────────────────────────────────────────────

@pytest.mark.parametrize("valor,esperado_prefixo", [
    ("=SOMASE(A1:A10)", "'"),
    ("+ATAQUES", "'"),
    ("-1234", "'"),
    ("@usuario", "'"),
    ("\t tab", "'"),
    ("\r cr", "'"),
])
def test_sanitizar_csv_prefixo_injecao(valor, esperado_prefixo):
    resultado = _sanitizar_csv(valor)
    assert resultado.startswith(esperado_prefixo), f"'{valor}' deveria receber prefixo '{esperado_prefixo}'"


@pytest.mark.parametrize("valor", [
    "Texto normal",
    "Governo anuncia medida",
    "123",
    "",
])
def test_sanitizar_csv_sem_prefixo(valor):
    resultado = _sanitizar_csv(valor)
    assert resultado == valor


def _criar_item(
    id=1,
    texto="Alegação de teste",
    tipo="texto",
    resultado="REAL",
    confianca=0.85,
    decisao_origem=None,
    justificativa=None,
    fontes=None,
):
    item = MagicMock()
    item.id = id
    item.texto_verificado = texto
    item.tipo = tipo
    item.resultado = resultado
    item.confianca = confianca
    item.criado_em = datetime(2026, 6, 25, 12, 0, 0)
    item.decisao_origem = decisao_origem
    item.justificativa_decisao = justificativa
    item.fontes_json = json.dumps(fontes) if fontes else None
    return item


# ── gerar_csv ─────────────────────────────────────────────────────────────────

def test_gerar_csv_cabecalho():
    csv_text = gerar_csv([])
    assert "ID" in csv_text
    assert "Alegação" in csv_text
    assert "Veredito" in csv_text


def test_gerar_csv_resultado_em_portugues():
    item = _criar_item(resultado="REAL")
    csv_text = gerar_csv([item])
    assert "Verdadeira" in csv_text
    assert "REAL" not in csv_text


def test_gerar_csv_previne_injecao():
    item = _criar_item(texto="=cmd(malicioso)")
    csv_text = gerar_csv([item])
    assert "'=cmd(malicioso)" in csv_text


# ── gerar_xlsx ────────────────────────────────────────────────────────────────

def test_gerar_xlsx_retorna_bytes():
    item = _criar_item()
    buf = gerar_xlsx([item])
    assert isinstance(buf, io.BytesIO)
    assert buf.read(4) == b"PK\x03\x04"  # ZIP magic do xlsx


def test_gerar_xlsx_registro_antigo_sem_decisao():
    """Registros sem decisao_origem devem mostrar mensagem de fallback."""
    item = _criar_item(decisao_origem=None, justificativa=None)
    buf = gerar_xlsx([item])

    wb = openpyxl.load_workbook(buf)
    ws = wb["Verificações"]

    valores = [str(c.value or "") for row in ws.iter_rows(min_row=2) for c in row]
    assert any(_NAO_DISPONIVEL in v for v in valores)


def test_gerar_xlsx_url_como_hyperlink():
    """URLs na aba de fontes devem ser armazenadas como hyperlinks clicáveis."""
    url = "https://g1.globo.com/noticia"
    item = _criar_item(fontes=[{
        "titulo": "Artigo G1",
        "url": url,
        "snippet": "Conteúdo...",
        "fonte": "g1.globo.com",
        "tipo_fonte": "jornalistica",
        "confiabilidade_fonte": "alta",
        "nli_label": "SUPPORTS",
        "nli_score": 0.9,
    }])
    buf = gerar_xlsx([item])

    wb = openpyxl.load_workbook(buf)
    ws = wb["Fontes"]

    # Verifica que a célula da URL (coluna 6) tem hyperlink
    url_cell = ws.cell(row=2, column=6)
    assert url_cell.hyperlink is not None
    assert str(url_cell.hyperlink.target) == url
