# ==============================================================================
# CheckAI API — Service: Verificações
# ==============================================================================

import csv
import io
import json
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from config import EXTRAIR_ARTIGOS_MAX
from db_models.verificacao import Verificacao
from models.schemas import VerificacaoResponse
from services.analisador_imagem import extrair_texto_imagem, extrair_texto_imagem_bytes
from services.busca_web import buscar_fontes
from services.classificador import classificar_texto
from services.concordancia_fontes import CONFIANCA_ALINHAMENTO_FRACO, reconciliar_com_fontes
from services.decisor_veredito import decidir_veredito_final
from services.extrator_artigos import extrair_conteudo_multiplos
from services.nli import agregar_resultado_nli, classificar_pares_nli
from services.ranking_fontes import ranquear_fontes


logger = logging.getLogger(__name__)

LIMIAR_INCONCLUSIVO = 0.60

# ---------------------------------------------------------------------------
# Exceção tipada para erros de link
# ---------------------------------------------------------------------------

class LinkError(Exception):
    """Erros de extração de link com tipo discriminado."""

    TIPO_URL_INVALIDA = "URL_INVALIDA"
    TIPO_PAGINA_INACESSIVEL = "PAGINA_INACESSIVEL"
    TIPO_CONTEUDO_NAO_ENCONTRADO = "CONTEUDO_NAO_ENCONTRADO"
    TIPO_CONTEUDO_INSUFICIENTE = "CONTEUDO_INSUFICIENTE"
    TIPO_TIMEOUT = "TIMEOUT"
    TIPO_ERRO_INTERNO = "ERRO_INTERNO"

    MENSAGENS_USUARIO = {
        TIPO_URL_INVALIDA: (
            "A URL informada é inválida. Verifique se ela começa com http:// ou https:// "
            "e possui um domínio válido."
        ),
        TIPO_PAGINA_INACESSIVEL: (
            "Não foi possível acessar a página. O site pode estar fora do ar, "
            "bloqueando acesso automatizado ou exigindo JavaScript."
        ),
        TIPO_CONTEUDO_NAO_ENCONTRADO: (
            "A página foi acessada, mas nenhum conteúdo principal foi identificado. "
            "Sites com paywall, login obrigatório ou JavaScript pesado podem causar isso."
        ),
        TIPO_CONTEUDO_INSUFICIENTE: (
            "O conteúdo extraído da página é muito curto para ser analisado. "
            "Tente informar o trecho da notícia diretamente no campo de texto."
        ),
        TIPO_TIMEOUT: (
            "A página demorou demais para responder. "
            "Tente novamente ou cole o texto da notícia diretamente."
        ),
        TIPO_ERRO_INTERNO: (
            "Ocorreu um erro ao processar a URL. "
            "Tente novamente ou cole o conteúdo diretamente como texto."
        ),
    }

    def __init__(self, tipo: str):
        self.tipo = tipo
        self.mensagem_usuario = self.MENSAGENS_USUARIO.get(tipo, "Erro ao processar a URL.")
        super().__init__(self.mensagem_usuario)


# ---------------------------------------------------------------------------
# Validação e extração de link
# ---------------------------------------------------------------------------

def _validar_url(url: str) -> None:
    """Levanta LinkError.TIPO_URL_INVALIDA se a URL não for válida."""
    url = url.strip()
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise LinkError(LinkError.TIPO_URL_INVALIDA)
    if not parsed.netloc or "." not in parsed.netloc:
        raise LinkError(LinkError.TIPO_URL_INVALIDA)


def _extrair_conteudo_link(url: str) -> dict:
    """
    Tenta extrair texto principal de uma URL com trafilatura.

    Retorna dict com: titulo, texto, dominio.
    Levanta LinkError com tipo adequado se falhar.
    """
    _validar_url(url)
    parsed = urlparse(url)
    dominio = parsed.netloc.lower().lstrip("www.")

    try:
        import trafilatura
        downloaded = trafilatura.fetch_url(url, no_ssl=True)
    except Exception as exc:
        msg = str(exc).lower()
        if "timeout" in msg or "timed out" in msg:
            raise LinkError(LinkError.TIPO_TIMEOUT)
        raise LinkError(LinkError.TIPO_PAGINA_INACESSIVEL)

    if not downloaded:
        raise LinkError(LinkError.TIPO_PAGINA_INACESSIVEL)

    try:
        meta = trafilatura.extract_metadata(downloaded)
        titulo = (meta.title or "").strip() if meta else ""

        resultado = trafilatura.extract(
            downloaded,
            include_comments=False,
            include_tables=False,
            favor_precision=True,
        )
    except Exception:
        raise LinkError(LinkError.TIPO_ERRO_INTERNO)

    if not resultado:
        raise LinkError(LinkError.TIPO_CONTEUDO_NAO_ENCONTRADO)

    texto = resultado.strip()
    if len(texto) < 80:
        raise LinkError(LinkError.TIPO_CONTEUDO_INSUFICIENTE)

    return {
        "titulo": titulo or dominio,
        "texto": texto,
        "dominio": dominio,
    }


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _mapear_resultado(classificacao: str, confianca: float) -> str:
    if confianca < LIMIAR_INCONCLUSIVO:
        return "INCONCLUSIVO"
    if classificacao == "VERDADEIRO":
        return "REAL"
    if classificacao == "FALSO":
        return "FALSO"
    return "INCONCLUSIVO"


def _enriquecer_fontes_com_artigos(fontes: list[dict]) -> None:
    if EXTRAIR_ARTIGOS_MAX <= 0 or not fontes:
        return
    urls = [f["url"] for f in fontes if f.get("url")]
    artigos_extraidos = extrair_conteudo_multiplos(
        urls, max_fontes=min(EXTRAIR_ARTIGOS_MAX, 3)
    )
    artigos_por_url = {a["url"]: a for a in artigos_extraidos}
    for fonte in fontes:
        artigo = artigos_por_url.get(fonte.get("url", ""))
        if artigo:
            fonte["texto_extraido"] = artigo.get("resumo", "")


_NLI_VOTOS_VAZIO: dict = {"SUPPORTS": 0, "REFUTES": 0, "NEUTRAL": 0}


def _enriquecer_fontes_com_nli(texto: str, fontes: list[dict]) -> tuple[str, float, dict]:
    try:
        nli_resultados = classificar_pares_nli(texto, fontes)
        for fonte, nli in zip(fontes, nli_resultados):
            fonte["nli_label"] = nli.get("label")
            fonte["nli_score"] = nli.get("score")
        agregado = agregar_resultado_nli(nli_resultados)
        return (
            agregado.get("label", "NEUTRAL"),
            float(agregado.get("score", 0.0)),
            agregado.get("votos", dict(_NLI_VOTOS_VAZIO)),
        )
    except Exception as exc:
        logger.warning("NLI falhou durante verificação, seguindo sem NLI: %s", exc)
        return "NEUTRAL", 0.0, dict(_NLI_VOTOS_VAZIO)


# ---------------------------------------------------------------------------
# Pipeline principal de verificação
# ---------------------------------------------------------------------------

def executar_verificacao(texto: str, tipo: str = "texto") -> dict:
    """
    Executa busca web, classificação ML e mapeamento de resultado.
    Para tipo="link" extrai o conteúdo da URL; levanta LinkError se falhar.
    """
    texto_para_analisar = texto
    texto_legivel = texto
    fonte_original: dict | None = None

    if tipo == "link":
        extraido = _extrair_conteudo_link(texto)
        texto_para_analisar = extraido["texto"]
        texto_legivel = extraido["titulo"]
        # Registra o URL original como fonte zero — será prependado às fontes externas
        fonte_original = {
            "titulo": extraido["titulo"],
            "url": texto,
            "snippet": extraido["texto"][:250],
            "fonte": extraido["dominio"],
            "tipo_fonte": "jornalistica",
            "confiabilidade_fonte": None,
            "peso_fonte": None,
            "nli_label": None,
            "nli_score": None,
            "texto_extraido": None,
        }
        logger.info("Link: extraídos %d chars (título: %s)", len(texto_para_analisar), texto_legivel)

    elif tipo == "imagem":
        texto_ocr = extrair_texto_imagem(texto)
        if texto_ocr:
            texto_para_analisar = texto_ocr
            logger.info("Imagem: OCR extraiu %d chars", len(texto_para_analisar))
        else:
            logger.warning("Imagem: OCR falhou — sem texto extraível na imagem")
            texto_para_analisar = ""
        texto_legivel = (
            "[Imagem enviada para análise]" if texto.startswith("data:") else texto
        )

    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_fontes = pool.submit(buscar_fontes, texto_para_analisar)
        fut_ml = pool.submit(classificar_texto, texto_para_analisar)
        fontes_externas = fut_fontes.result()
        resultado_ml = fut_ml.result()

    fontes = ranquear_fontes(fontes_externas, max_fontes=5)
    _enriquecer_fontes_com_artigos(fontes)

    # Para links sem evidências externas, o sistema não pode confirmar a alegação
    if tipo == "link" and not fontes:
        if fonte_original:
            fontes = [fonte_original]
        return {
            "texto_verificado": texto_legivel,
            "tipo": tipo,
            "resultado": "INCONCLUSIVO",
            "confianca": CONFIANCA_ALINHAMENTO_FRACO,
            "modelo_ativo": "sim" if resultado_ml["modelo_ativo"] else "nao",
            "fontes": fontes,
            "nli_resultado_agregado": "NEUTRAL",
            "nli_score_agregado": 0.0,
            "nli_votos": dict(_NLI_VOTOS_VAZIO),
            "decisao_origem": "sem_evidencias_externas",
            "justificativa_decisao": (
                "Não foram encontradas fontes externas para verificar as informações "
                "deste link. Sem evidências adicionais, não é possível confirmar nem "
                "refutar a alegação."
            ),
        }

    # Prepend fonte_original (URL enviada) no início da lista de fontes
    if fonte_original:
        urls_ja_presentes = {f.get("url", "") for f in fontes}
        if texto not in urls_ja_presentes:
            fontes.insert(0, fonte_original)

    nli_resultado_agregado, nli_score_agregado, nli_votos = _enriquecer_fontes_com_nli(
        texto_para_analisar, fontes
    )

    classificacao, confianca, _ = reconciliar_com_fontes(
        texto_para_analisar,
        resultado_ml["classificacao"],
        resultado_ml["confianca"],
        fontes,
    )

    resultado = _mapear_resultado(classificacao, confianca)
    if resultado == "INCONCLUSIVO" and confianca < CONFIANCA_ALINHAMENTO_FRACO:
        confianca = CONFIANCA_ALINHAMENTO_FRACO

    veredito = decidir_veredito_final(
        resultado_atual=resultado,
        confianca_atual=confianca,
        nli_resultado_agregado=nli_resultado_agregado,
        nli_score_agregado=nli_score_agregado,
        nli_votos=nli_votos,
        fontes=fontes,
    )
    resultado = veredito["resultado"]
    confianca = veredito["confianca"]

    return {
        "texto_verificado": texto_legivel,
        "tipo": tipo,
        "resultado": resultado,
        "confianca": confianca,
        "modelo_ativo": "sim" if resultado_ml["modelo_ativo"] else "nao",
        "fontes": fontes,
        "nli_resultado_agregado": nli_resultado_agregado,
        "nli_score_agregado": nli_score_agregado,
        "nli_votos": nli_votos,
        "decisao_origem": veredito["decisao_origem"],
        "justificativa_decisao": veredito["justificativa_decisao"],
    }


def executar_verificacao_imagem(conteudo: bytes, nome_arquivo: str | None = None) -> dict:
    nome_log = nome_arquivo or "<sem nome>"
    logger.info("OCR: processando '%s' (%d bytes)", nome_log, len(conteudo))

    texto_ocr = extrair_texto_imagem_bytes(conteudo)

    if not texto_ocr:
        logger.warning("OCR: nenhum texto legível extraído de '%s'", nome_log)
        justificativa = "Não foi possível identificar texto legível na imagem enviada."
        return {
            "texto_verificado": justificativa,
            "tipo": "imagem",
            "resultado": "INCONCLUSIVO",
            "confianca": CONFIANCA_ALINHAMENTO_FRACO,
            "modelo_ativo": "nao",
            "fontes": [],
            "nli_resultado_agregado": None,
            "nli_score_agregado": None,
            "nli_votos": None,
            "decisao_origem": None,
            "justificativa_decisao": justificativa,
        }

    logger.info("OCR: extraídos %d chars de '%s'", len(texto_ocr), nome_log)

    with ThreadPoolExecutor(max_workers=2) as pool:
        fut_fontes = pool.submit(buscar_fontes, texto_ocr)
        fut_ml = pool.submit(classificar_texto, texto_ocr)
        fontes = fut_fontes.result()
        resultado_ml = fut_ml.result()

    fontes = ranquear_fontes(fontes, max_fontes=5)
    _enriquecer_fontes_com_artigos(fontes)

    nli_resultado_agregado, nli_score_agregado, nli_votos = _enriquecer_fontes_com_nli(
        texto_ocr, fontes
    )

    classificacao, confianca, _ = reconciliar_com_fontes(
        texto_ocr,
        resultado_ml["classificacao"],
        resultado_ml["confianca"],
        fontes,
    )

    resultado = _mapear_resultado(classificacao, confianca)
    if resultado == "INCONCLUSIVO" and confianca < CONFIANCA_ALINHAMENTO_FRACO:
        confianca = CONFIANCA_ALINHAMENTO_FRACO

    veredito = decidir_veredito_final(
        resultado_atual=resultado,
        confianca_atual=confianca,
        nli_resultado_agregado=nli_resultado_agregado,
        nli_score_agregado=nli_score_agregado,
        nli_votos=nli_votos,
        fontes=fontes,
    )

    return {
        "texto_verificado": texto_ocr,
        "tipo": "imagem",
        "resultado": veredito["resultado"],
        "confianca": veredito["confianca"],
        "modelo_ativo": "sim" if resultado_ml["modelo_ativo"] else "nao",
        "fontes": fontes,
        "nli_resultado_agregado": nli_resultado_agregado,
        "nli_score_agregado": nli_score_agregado,
        "nli_votos": nli_votos,
        "decisao_origem": veredito["decisao_origem"],
        "justificativa_decisao": veredito["justificativa_decisao"],
    }


# ---------------------------------------------------------------------------
# CRUD — persistência
# ---------------------------------------------------------------------------

def _dados_para_verificacao(dados: dict) -> dict:
    """Extrai apenas os campos que existem no ORM Verificacao."""
    return {
        "texto_verificado": dados["texto_verificado"],
        "tipo": dados["tipo"],
        "resultado": dados["resultado"],
        "confianca": dados["confianca"],
        "modelo_ativo": dados["modelo_ativo"],
        "fontes_json": (
            json.dumps(dados["fontes"], ensure_ascii=False) if dados.get("fontes") else None
        ),
        "nli_resultado_agregado": dados.get("nli_resultado_agregado"),
        "nli_score_agregado": dados.get("nli_score_agregado"),
        "nli_votos_json": (
            json.dumps(dados["nli_votos"], ensure_ascii=False)
            if dados.get("nli_votos")
            else None
        ),
        "decisao_origem": dados.get("decisao_origem"),
        "justificativa_decisao": dados.get("justificativa_decisao"),
    }


def criar_verificacao_por_imagem(
    db: Session, usuario_id: int, conteudo: bytes, nome_arquivo: str | None = None
) -> Verificacao:
    dados = executar_verificacao_imagem(conteudo, nome_arquivo)
    verificacao = Verificacao(usuario_id=usuario_id, **_dados_para_verificacao(dados))
    db.add(verificacao)
    db.commit()
    db.refresh(verificacao)
    return verificacao


def criar_verificacao(
    db: Session, usuario_id: int, texto: str, tipo: str = "texto"
) -> Verificacao:
    dados = executar_verificacao(texto, tipo)
    verificacao = Verificacao(usuario_id=usuario_id, **_dados_para_verificacao(dados))
    db.add(verificacao)
    db.commit()
    db.refresh(verificacao)
    return verificacao


def verificacao_para_response(verificacao: Verificacao) -> VerificacaoResponse:
    return VerificacaoResponse.model_validate(verificacao)


# ---------------------------------------------------------------------------
# Listagem e filtros
# ---------------------------------------------------------------------------

def listar_verificacoes(
    db: Session,
    usuario_id: int,
    *,
    resultado: str | None = None,
    busca: str | None = None,
    tipo: str | None = None,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
    ordenacao: str = "desc",
    pagina: int = 1,
    por_pagina: int = 10,
) -> dict:
    """Lista o histórico do usuário com filtros e paginação."""
    query = db.query(Verificacao).filter(Verificacao.usuario_id == usuario_id)

    if resultado:
        query = query.filter(Verificacao.resultado == resultado.upper())

    if tipo:
        query = query.filter(Verificacao.tipo == tipo.lower())

    if busca:
        query = query.filter(Verificacao.texto_verificado.ilike(f"%{busca}%"))

    if data_inicio:
        query = query.filter(Verificacao.criado_em >= data_inicio)

    if data_fim:
        # Inclui todo o dia final (até 23:59:59)
        fim_dia = data_fim.replace(hour=23, minute=59, second=59)
        query = query.filter(Verificacao.criado_em <= fim_dia)

    total = query.count()

    order_col = (
        Verificacao.criado_em.asc()
        if ordenacao == "asc"
        else Verificacao.criado_em.desc()
    )
    offset = (pagina - 1) * por_pagina
    itens = query.order_by(order_col).offset(offset).limit(por_pagina).all()

    total_paginas = (total + por_pagina - 1) // por_pagina if total > 0 else 0

    return {
        "total": total,
        "pagina": pagina,
        "por_pagina": por_pagina,
        "total_paginas": total_paginas,
        "itens": itens,
    }


def buscar_verificacao_por_id(
    db: Session, usuario_id: int, verificacao_id: int
) -> Verificacao | None:
    return (
        db.query(Verificacao)
        .filter(
            Verificacao.id == verificacao_id,
            Verificacao.usuario_id == usuario_id,
        )
        .first()
    )


# ---------------------------------------------------------------------------
# Resumo / Dashboard
# ---------------------------------------------------------------------------

def _query_resumo(db: Session, usuario_id: int, data_inicio=None, data_fim=None):
    query = db.query(Verificacao).filter(Verificacao.usuario_id == usuario_id)
    if data_inicio:
        query = query.filter(Verificacao.criado_em >= data_inicio)
    if data_fim:
        query = query.filter(Verificacao.criado_em <= data_fim)
    return query


def calcular_resumo(db: Session, usuario_id: int) -> dict:
    return calcular_resumo_periodo(db, usuario_id)


def calcular_resumo_periodo(
    db: Session,
    usuario_id: int,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
) -> dict:
    verificacoes = _query_resumo(db, usuario_id, data_inicio, data_fim).all()
    total = len(verificacoes)
    reais = sum(1 for v in verificacoes if v.resultado == "REAL")
    falsas = sum(1 for v in verificacoes if v.resultado == "FALSO")
    inconclusivas = sum(1 for v in verificacoes if v.resultado == "INCONCLUSIVO")
    percentual = round((reais / total * 100), 1) if total > 0 else 0.0
    return {
        "total_verificacoes": total,
        "total_reais": reais,
        "total_falsas": falsas,
        "total_inconclusivas": inconclusivas,
        "percentual_reais": percentual,
    }


def periodo_para_datas(periodo: str) -> tuple[datetime | None, datetime | None]:
    """Converte string de período para intervalo de datas UTC."""
    agora = datetime.utcnow()
    mapa = {"7d": 7, "30d": 30, "90d": 90}
    dias = mapa.get(periodo)
    if dias:
        return agora - timedelta(days=dias), agora
    return None, None  # "all"


# ---------------------------------------------------------------------------
# Fontes mais consultadas
# ---------------------------------------------------------------------------

def listar_fontes_top(
    db: Session,
    usuario_id: int,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
    limite: int = 5,
) -> list[dict]:
    """
    Agrega os domínios mais consultados nas verificações do usuário,
    parseando fontes_json de cada registro persistido.
    """
    query = db.query(Verificacao).filter(
        Verificacao.usuario_id == usuario_id,
        Verificacao.fontes_json.isnot(None),
    )
    if data_inicio:
        query = query.filter(Verificacao.criado_em >= data_inicio)
    if data_fim:
        query = query.filter(Verificacao.criado_em <= data_fim)

    verificacoes = query.all()

    dominio_info: dict[str, dict] = {}
    total_fontes = 0

    for v in verificacoes:
        try:
            fontes: list[dict] = json.loads(v.fontes_json) if v.fontes_json else []
        except (json.JSONDecodeError, TypeError):
            fontes = []

        for fonte in fontes:
            dominio = fonte.get("fonte", "").strip()
            if not dominio:
                try:
                    dominio = urlparse(fonte.get("url", "")).netloc.lstrip("www.").strip()
                except Exception:
                    pass
            if not dominio:
                continue

            total_fontes += 1
            if dominio not in dominio_info:
                dominio_info[dominio] = {
                    "dominio": dominio,
                    "total": 0,
                    "tipo_fonte": fonte.get("tipo_fonte"),
                    "confiabilidade": fonte.get("confiabilidade_fonte"),
                }
            dominio_info[dominio]["total"] += 1

    top = sorted(dominio_info.values(), key=lambda x: x["total"], reverse=True)[:limite]

    for item in top:
        item["percentual"] = (
            round(item["total"] / total_fontes * 100, 1) if total_fontes > 0 else 0.0
        )

    return top


# ---------------------------------------------------------------------------
# Exportação
# ---------------------------------------------------------------------------

_RESULTADO_MAP = {"REAL": "Verdadeira", "FALSO": "Falsa", "INCONCLUSIVO": "Inconclusivo"}
_TIPO_MAP = {"texto": "Texto", "link": "Link", "imagem": "Imagem"}
_NAO_DISPONIVEL = "Não disponível — registro anterior à atualização"

_HEADER_ESTILO_FUNDO = "1E293B"
_HEADER_ESTILO_FONTE = "FFFFFF"


def _sanitizar_csv(valor) -> str:
    """Previne CSV injection para valores iniciados com =, +, -, @, tab ou CR."""
    s = str(valor) if valor is not None else ""
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


def _estilizar_cabecalho(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color=_HEADER_ESTILO_FONTE)
        cell.fill = PatternFill(
            start_color=_HEADER_ESTILO_FUNDO,
            end_color=_HEADER_ESTILO_FUNDO,
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")


def gerar_xlsx(itens: list[Verificacao]) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # ── Aba 1: Verificações ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Verificações"

    cabecalhos_v = [
        "ID", "Data e Hora", "Alegação", "Tipo",
        "Veredito", "Confiança (%)", "Origem da Decisão",
        "Justificativa", "Qtd. Fontes",
    ]
    ws1.append(cabecalhos_v)
    _estilizar_cabecalho(ws1)

    for item in itens:
        try:
            fontes = json.loads(item.fontes_json) if item.fontes_json else []
        except (json.JSONDecodeError, TypeError):
            fontes = []

        ws1.append([
            item.id,
            item.criado_em.strftime("%d/%m/%Y %H:%M:%S"),
            item.texto_verificado,
            _TIPO_MAP.get(item.tipo, item.tipo),
            _RESULTADO_MAP.get(item.resultado, item.resultado),
            round(item.confianca * 100, 1),
            item.decisao_origem or _NAO_DISPONIVEL,
            item.justificativa_decisao or _NAO_DISPONIVEL,
            len(fontes),
        ])

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    larguras_v = [8, 18, 60, 10, 15, 12, 28, 55, 12]
    for i, w in enumerate(larguras_v, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Aba 2: Fontes ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Fontes")

    cabecalhos_f = [
        "ID Verificação", "Data Verificação", "Alegação",
        "Título da Fonte", "Domínio", "URL",
        "Tipo da Fonte", "Confiabilidade", "NLI", "Score NLI",
    ]
    ws2.append(cabecalhos_f)
    _estilizar_cabecalho(ws2)

    for item in itens:
        try:
            fontes = json.loads(item.fontes_json) if item.fontes_json else []
        except (json.JSONDecodeError, TypeError):
            fontes = []

        data_str = item.criado_em.strftime("%d/%m/%Y %H:%M:%S")
        for fonte in fontes:
            score = fonte.get("nli_score")
            url_val = fonte.get("url", "")
            ws2.append([
                item.id,
                data_str,
                item.texto_verificado,
                fonte.get("titulo", ""),
                fonte.get("fonte", ""),
                url_val,
                fonte.get("tipo_fonte", ""),
                fonte.get("confiabilidade_fonte", ""),
                fonte.get("nli_label", ""),
                round(score * 100, 1) if score is not None else "",
            ])
            # Torna a URL clicável no Excel
            if url_val:
                cell = ws2.cell(row=ws2.max_row, column=6)
                cell.hyperlink = url_val
                cell.style = "Hyperlink"

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    larguras_f = [15, 18, 45, 45, 28, 65, 15, 15, 12, 10]
    for i, w in enumerate(larguras_f, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_csv(itens: list[Verificacao]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    writer.writerow([
        "ID", "Data e Hora", "Alegação", "Tipo",
        "Veredito", "Confiança (%)", "Origem da Decisão",
        "Justificativa", "Qtd. Fontes",
    ])

    for item in itens:
        try:
            fontes = json.loads(item.fontes_json) if item.fontes_json else []
        except (json.JSONDecodeError, TypeError):
            fontes = []

        writer.writerow([
            _sanitizar_csv(item.id),
            _sanitizar_csv(item.criado_em.strftime("%d/%m/%Y %H:%M:%S")),
            _sanitizar_csv(item.texto_verificado),
            _sanitizar_csv(_TIPO_MAP.get(item.tipo, item.tipo)),
            _sanitizar_csv(_RESULTADO_MAP.get(item.resultado, item.resultado)),
            _sanitizar_csv(round(item.confianca * 100, 1)),
            _sanitizar_csv(item.decisao_origem or ""),
            _sanitizar_csv(item.justificativa_decisao or ""),
            _sanitizar_csv(len(fontes)),
        ])

    return output.getvalue()


# ---------------------------------------------------------------------------
# Série temporal para gráfico de barras do dashboard
# ---------------------------------------------------------------------------

_MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
             "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def obter_serie_temporal(
    db: Session,
    usuario_id: int,
    data_inicio: datetime | None,
    data_fim: datetime | None,
    periodo: str,
) -> list[dict]:
    """
    Retorna verificações agrupadas por intervalo temporal.
    - 7d / 30d → por dia  (label = "DD/MM")
    - 90d       → por semana (label = "SNN")
    - all       → por mês  (label = "Mmm/AA")
    """
    query = db.query(Verificacao).filter(Verificacao.usuario_id == usuario_id)
    if data_inicio:
        query = query.filter(Verificacao.criado_em >= data_inicio)
    if data_fim:
        query = query.filter(Verificacao.criado_em <= data_fim)

    verificacoes = query.order_by(Verificacao.criado_em.asc()).all()

    if periodo in ("7d", "30d"):
        def get_key(v: Verificacao) -> str:
            return v.criado_em.strftime("%Y-%m-%d")

        def get_label(key: str) -> str:
            return datetime.strptime(key, "%Y-%m-%d").strftime("%d/%m")

    elif periodo == "90d":
        def get_key(v: Verificacao) -> str:
            return v.criado_em.strftime("%G-%V")  # ISO year-week

        def get_label(key: str) -> str:
            return f"S{int(key.split('-')[1])}"

    else:  # all
        def get_key(v: Verificacao) -> str:
            return v.criado_em.strftime("%Y-%m")

        def get_label(key: str) -> str:
            year, month = key.split("-")
            return f"{_MESES_PT[int(month) - 1]}/{int(year) % 100:02d}"

    groups: dict[str, dict] = {}
    for v in verificacoes:
        key = get_key(v)
        if key not in groups:
            groups[key] = {"key": key, "label": get_label(key),
                           "verdadeiras": 0, "falsas": 0, "inconclusivas": 0}
        if v.resultado == "REAL":
            groups[key]["verdadeiras"] += 1
        elif v.resultado == "FALSO":
            groups[key]["falsas"] += 1
        else:
            groups[key]["inconclusivas"] += 1

    # Preenche slots vazios para 7d e 30d (garante todos os dias no período)
    if periodo in ("7d", "30d") and data_inicio:
        dias = 7 if periodo == "7d" else 30
        for i in range(dias):
            d = data_inicio + timedelta(days=i)
            key = d.strftime("%Y-%m-%d")
            if key not in groups:
                groups[key] = {
                    "key": key, "label": d.strftime("%d/%m"),
                    "verdadeiras": 0, "falsas": 0, "inconclusivas": 0,
                }

    sorted_groups = sorted(groups.values(), key=lambda x: x["key"])
    return [
        {
            "label": g["label"],
            "verdadeiras": g["verdadeiras"],
            "falsas": g["falsas"],
            "inconclusivas": g["inconclusivas"],
        }
        for g in sorted_groups
    ]


# ---------------------------------------------------------------------------
# Limpeza de histórico
# ---------------------------------------------------------------------------

def limpar_historico(db: Session, usuario_id: int) -> int:
    total = (
        db.query(Verificacao)
        .filter(Verificacao.usuario_id == usuario_id)
        .delete(synchronize_session=False)
    )
    db.commit()
    return total
